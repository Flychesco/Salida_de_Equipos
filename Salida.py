import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, simpledialog
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
import subprocess
import win32com.client
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

class ExcelController:
    def __init__(self, master):
        self.master = master
        master.title("SALIDAS GAME")
        
        # Aumentar el tamaño de la ventana y centrarla en la pantalla
        master.geometry("800x600")
        master.update_idletasks()
        width = master.winfo_width()
        height = master.winfo_height()
        x = (master.winfo_screenwidth() // 2) - (width // 2)
        y = (master.winfo_screenheight() // 2) - (height // 2)
        master.geometry('{}x{}+{}+{}'.format(width, height, x, y))

        # Frame principal para centrar los widgets
        main_frame = tk.Frame(master)
        main_frame.place(relx=0.5, rely=0.5, anchor='center')

        # Variables
        self.reference = tk.StringVar()
        self.value1 = tk.StringVar()
        self.value2 = tk.StringVar()
        self.value3 = tk.StringVar()
        self.value4 = tk.StringVar()
        self.excel_files = {
            "H": "Salida Equipos H.xlsx",
            "V": "Salida Equipos V.xlsx"
        }

        # Crear widgets
        tk.Label(main_frame, text="Equipo:").grid(row=0, column=0, sticky="e", padx=10, pady=5)
        self.entry_reference = tk.Entry(main_frame, textvariable=self.reference, validate="key", width=20)
        self.entry_reference['validatecommand'] = (self.entry_reference.register(self.validate_numeric_input), '%P', 6)
        self.entry_reference.grid(row=0, column=1, padx=10, pady=5)

        tk.Label(main_frame, text="SKU:").grid(row=1, column=0, sticky="e", padx=10, pady=5)
        self.entry_value1 = tk.Entry(main_frame, textvariable=self.value1, state="disabled", validate="key", width=20)
        self.entry_value1['validatecommand'] = (self.entry_value1.register(self.validate_numeric_input), '%P', 6)
        self.entry_value1.grid(row=1, column=1, padx=10, pady=5)

        tk.Label(main_frame, text="RMA:").grid(row=2, column=0, sticky="e", padx=10, pady=5)
        self.entry_value2 = tk.Entry(main_frame, textvariable=self.value2, state="disabled", width=20)
        self.entry_value2.grid(row=2, column=1, padx=10, pady=5)

        tk.Label(main_frame, text="$:").grid(row=3, column=0, sticky="e", padx=10, pady=5)
        self.entry_value3 = tk.Entry(main_frame, textvariable=self.value3, state="disabled", width=20)
        self.entry_value3.grid(row=3, column=1, padx=10, pady=5)

        tk.Label(main_frame, text="Trabajo_Realizado:").grid(row=4, column=0, sticky="ne", padx=10, pady=5)
        self.entry_value4 = scrolledtext.ScrolledText(main_frame, width=60, height=4, wrap=tk.WORD, state="disabled")
        self.entry_value4.grid(row=4, column=1, padx=10, pady=5)

        self.btn_unlock = tk.Button(main_frame, text="Buscar Equipo", command=self.unlock_fields, width=15)
        self.btn_unlock.grid(row=5, column=0, columnspan=2, pady=10)

        button_frame = tk.Frame(main_frame)
        button_frame.grid(row=6, column=0, columnspan=2, pady=10)

        self.btn_save_h = tk.Button(button_frame, text="Guardar en H", command=lambda: self.save_to_excel("H"), width=15)
        self.btn_save_h.pack(side=tk.LEFT, padx=5)

        self.btn_save_v = tk.Button(button_frame, text="Guardar en V", command=lambda: self.save_to_excel("V"), width=15)
        self.btn_save_v.pack(side=tk.LEFT, padx=5)

        self.btn_create_folder = tk.Button(main_frame, text="Guardar imagen", command=self.create_folder, width=15)
        self.btn_create_folder.grid(row=7, column=0, columnspan=2, pady=10)

        self.btn_generate_pdf = tk.Button(main_frame, text="Generar PDF", command=self.generate_pdf, width=15)
        self.btn_generate_pdf.grid(row=8, column=0, columnspan=2, pady=10)

        #botón "Crear salida"
        self.btn_create_output = tk.Button(main_frame, text="Crear salida", command=self.create_output, width=15)
        self.btn_create_output.grid(row=9, column=0, columnspan=2, pady=10)
    
    def validate_numeric_input(self, value, max_length):
        return len(value) <= int(max_length) and (value.isdigit() or value == "")

    def unlock_fields(self):
        if self.reference.get():
            self.entry_value1.config(state="normal")
            self.entry_value2.config(state="normal")
            self.entry_value3.config(state="normal")
            self.entry_value4.config(state="normal")
        else:
            messagebox.showerror("Error", "Por favor, ingrese una referencia")

    def create_output(self):
        if messagebox.askyesno("Confirmación", "¿Está seguro de que desea crear la salida?"):
            # Preguntar si es H o V
            choice = simpledialog.askstring("Selección", "¿Desea crear la salida para H o V?", initialvalue="H")
            if choice:
                choice = choice.upper()  # Convertir a mayúsculas
                if choice not in ["H", "V"]:
                    messagebox.showerror("Error", "Opción no válida. Por favor, elija H o V.")
                    return
            else:
                messagebox.showerror("Error", "Debe seleccionar H o V.")
                return

            try:
                # Crear instancia de Outlook
                outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
                
                # Crear un nuevo correo
                mail = outlook.CreateItem(0)
                
                # Configurar destinatarios (reemplaza con las direcciones reales)
                mail.To = "destinatario1@ejemplo.com; destinatario2@ejemplo.com"
                mail.CC = "copia1@ejemplo.com; copia2@ejemplo.com"
                
                # Configurar asunto y cuerpo del correo
                mail.Subject = f"Salida de equipos - {choice}"
                mail.Body = f"Adjunto encontrará el informe de salida de equipos {choice} y las fotografías correspondientes."
                
                # Adjuntar el archivo PDF
                pdf_file = f"Salida Equipos {choice}.pdf"
                if os.path.exists(pdf_file):
                    mail.Attachments.Add(os.path.abspath(pdf_file))
                else:
                    messagebox.showwarning("Advertencia", f"No se encontró el archivo PDF para {choice}.")
                
                # Adjuntar las fotografías
                photo_folder = "Fotos_Equipos"  # Reemplaza con la ruta real de la carpeta de fotos
                if os.path.exists(photo_folder):
                    wb = openpyxl.load_workbook(self.excel_files[choice])
                    ws = wb.active
                    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                        reference = row[0]
                        photo_path = os.path.join(photo_folder, f"{reference}.jpg")
                        if os.path.exists(photo_path):
                            mail.Attachments.Add(os.path.abspath(photo_path))
                        else:
                            messagebox.showwarning("Advertencia", f"No se encontró la foto para la referencia {reference}.")
                else:
                    messagebox.showwarning("Advertencia", f"No se encontró la carpeta de fotos: {photo_folder}")
                
                # Mostrar el correo (no lo envía automáticamente)
                mail.Display(True)
                
                messagebox.showinfo("Éxito", f"Correo creado con éxito para {choice}. Por favor, revíselo antes de enviar.")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo crear el correo: {str(e)}")

    def save_to_excel(self, excel_choice):
        if not all([self.reference.get(), self.value1.get(), self.value2.get(), self.value3.get(), self.entry_value4.get("1.0", tk.END).strip()]):
            messagebox.showerror("Error", "Por favor, complete todos los campos")
            return
        excel_file = self.excel_files[excel_choice]
        # Verificar si el archivo existe, si no, crearlo
        if not os.path.exists(excel_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Datos"
            headers = ["EQUIPO", "SKU", "Nº RMA", "PRECIO", "TRABAJO REALIZADO"]
            ws.append(headers)
            self.format_headers(ws)
        else:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

        # Contar el número de filas con datos
        row_count = sum(1 for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0])

        if row_count >= 36 and not self.reference_exists(ws):
            messagebox.showerror("Error", "Se ha alcanzado el límite de 36 entradas")
            return

        # Buscar si la referencia ya existe
        reference_exists = False
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] == self.reference.get():
                reference_exists = True
                break

        if reference_exists:
            # Actualizar la fila existente
            for row in ws.iter_rows(min_row=2):
                if row[0].value == self.reference.get():
                    row[1].value = self.value1.get()
                    row[2].value = self.value2.get()
                    row[3].value = self.value3.get()
                    row[4].value = self.entry_value4.get("1.0", tk.END).strip().upper()  # Convertir a mayúsculas
                    break
        else:
            # Añadir una nueva fila
            values = [self.reference.get(), self.value1.get(), self.value2.get(), self.value3.get(), self.entry_value4.get("1.0", tk.END).strip().upper()]
            ws.append(values)

        # Ajustar el ancho de las columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(excel_file)
        messagebox.showinfo("Éxito", f"Datos guardados correctamente en el archivo Excel {excel_choice}")
        
        # Resetear los campos después de guardar
        self.reset_fields()

    def reset_fields(self):
        self.reference.set("")
        self.value1.set("")
        self.value2.set("")
        self.value3.set("")
        self.entry_value4.delete("1.0", tk.END)
        self.entry_value1.config(state="disabled")
        self.entry_value2.config(state="disabled")
        self.entry_value3.config(state="disabled")
        self.entry_value4.config(state="disabled")
        
    def create_folder(self):
        if not self.reference.get():
            messagebox.showerror("Error", "Por favor, ingrese una referencia")
            return

        folder_name = self.reference.get()
        try:
            os.mkdir(folder_name)
            messagebox.showinfo("Éxito", f"Carpeta '{folder_name}' creada correctamente")
            
            # Abrir la carpeta automáticamente
            if os.name == 'nt':  # Para Windows
                os.startfile(folder_name)
            elif os.name == 'posix':  # Para macOS y Linux
                subprocess.call(('open', folder_name))
        except FileExistsError:
            messagebox.showerror("Error", f"La carpeta '{folder_name}' ya existe")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear la carpeta: {str(e)}")

    def format_headers(self, worksheet):
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def reference_exists(self, worksheet):
        for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] == self.reference.get():
                return True
        return False

    def generate_pdf(self):
        excel_choice = messagebox.askquestion("Seleccionar archivo", "¿Desea generar el PDF del archivo H?")
        excel_file = self.excel_files["H"] if excel_choice == 'yes' else self.excel_files["V"]
        
        if not os.path.exists(excel_file):
            messagebox.showerror("Error", f"El archivo {excel_file} no existe")
            return

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)

        pdf_file = f"{os.path.splitext(excel_file)[0]}.pdf"
        doc = SimpleDocTemplate(pdf_file, pagesize=letter)
        elements = []

        table = Table(data)
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        table.setStyle(style)
        elements.append(table)

        doc.build(elements)
        messagebox.showinfo("Éxito", f"PDF generado correctamente: {pdf_file}")

root = tk.Tk()
app = ExcelController(root)
root.mainloop()